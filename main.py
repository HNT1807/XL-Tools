
import streamlit as st
import io
import base64
from openpyxl import load_workbook, Workbook

st.set_page_config(page_title="XL TOOLS", layout="wide")

st.markdown("""
<style>
    .main-title { font-size: 50px; font-weight: bold; text-align: center; margin-bottom: 30px; }
    .stApp > header { display: none !important; }
    .block-container { max-width: 1100px; padding-top: 2rem; padding-bottom: 10rem; }
    .stButton {margin-right: 20px;}
    .delete-button {color: red; cursor: pointer;}
    /* Add styling for filter rows */
    .filter-row {
        margin-bottom: 1rem;
        padding: 0.5rem;
        border-radius: 0.3rem;
    }
</style>
""", unsafe_allow_html=True)

st.markdown("<h1 class='main-title'>XL TOOLS</h1>", unsafe_allow_html=True)

uploaded_files = st.file_uploader("Drag and drop Excel files here", type=["xlsx", "xls"], accept_multiple_files=True)

# Initialize session state for filters
if 'filters' not in st.session_state:
    st.session_state.filters = [{
        'enabled': False,
        'header': '',
        'operator': 'Containing',
        'value': '',
        'case_sensitive': 'Non Case Sensitive'
    }]


def add_filter():
    st.session_state.filters.append({
        'enabled': True,
        'header': '',
        'operator': 'Containing',
        'value': '',
        'case_sensitive': 'Non Case Sensitive'
    })


def remove_filter(index):
    if len(st.session_state.filters) > 1:  # Keep at least one filter row
        st.session_state.filters.pop(index)


def render_filter_row(idx):
    filter_data = st.session_state.filters[idx]

    col_enable, col_header, col_operator, col_value, col_case, col_delete = st.columns([0.5, 1, 1, 1, 1, 0.3])

    with col_enable:
        if idx == 0:
            st.write("")
            st.write("")
            enabled = st.checkbox("Filter", key=f"enable_{idx}")
        else:
            st.write("")
            st.write("")
            st.markdown("**AND**")
            enabled = True
        filter_data['enabled'] = enabled

    with col_header:
        filter_data['header'] = st.text_input(
            "",
            value=filter_data['header'],
            placeholder="Enter header name",
            key=f"header_{idx}",
            disabled=not (idx == 0 and enabled or idx > 0)
        )

    with col_operator:
        filter_data['operator'] = st.selectbox(
            "",
            ["Containing", "Don't Contain", "Is", "Is Not"],
            key=f"operator_{idx}",
            disabled=not (idx == 0 and enabled or idx > 0)
        )

    with col_value:
        filter_data['value'] = st.text_input(
            "",
            value=filter_data['value'],
            placeholder="Enter value",
            key=f"value_{idx}",
            disabled=not (idx == 0 and enabled or idx > 0)
        )

    with col_case:
        filter_data['case_sensitive'] = st.selectbox(
            "",
            ["Case Sensitive", "Non Case Sensitive"],
            key=f"case_{idx}",
            disabled=not (idx == 0 and enabled or idx > 0)
        )

    with col_delete:
        if idx > 0:  # Only show delete button for additional filters
            st.write("")
            st.write("")
            if st.button("🗑️", key=f"delete_{idx}"):
                remove_filter(idx)
                st.rerun()


# Render filter interface
st.markdown("<br>", unsafe_allow_html=True)
for i in range(len(st.session_state.filters)):
    render_filter_row(i)

# Add filter button
col1, col2, col3 = st.columns([1, 4, 1])
with col2:
    if st.button("➕ Add another filter"):
        add_filter()
        st.rerun()

st.markdown("<br>", unsafe_allow_html=True)


def should_include_row(row, header_indices, filters):
    """Check if row meets all filter criteria"""
    # If no filters are enabled, return True
    if not any(f['enabled'] for f in filters):
        return True

    # Check each enabled filter
    for idx, (header_idx, filter_config) in enumerate(zip(header_indices, filters)):
        if not filter_config['enabled']:
            continue

        if header_idx is None:
            continue

        cell_value = str(row[header_idx].value or "")
        filter_value = str(filter_config['value'])

        # Apply case sensitivity
        if filter_config['case_sensitive'] == "Non Case Sensitive":
            cell_value = cell_value.lower()
            filter_value = filter_value.lower()

        # Apply filter based on operator
        if filter_config['operator'] == "Containing":
            if filter_value not in cell_value:
                return False
        elif filter_config['operator'] == "Don't Contain":
            if filter_value in cell_value:
                return False
        elif filter_config['operator'] == "Is":
            if cell_value != filter_value:
                return False
        elif filter_config['operator'] == "Is Not":
            if cell_value == filter_value:
                return False

    return True


def find_header_index(sheet, header_name):
    """Find the column index for a given header name"""
    for row in sheet.iter_rows(min_row=1, max_row=1):
        for idx, cell in enumerate(row):
            if cell.value and str(cell.value).lower() == header_name.lower():
                return idx
    return None


def copy_column_dimensions(source_sheet, target_sheet):
    for column_letter, column_dimension in source_sheet.column_dimensions.items():
        target_sheet.column_dimensions[column_letter].width = column_dimension.width


def copy_row_dimensions(source_sheet, target_sheet):
    for row_idx, row_dimension in source_sheet.row_dimensions.items():
        target_sheet.row_dimensions[row_idx].height = row_dimension.height


def copy_row_to_sheet(source_row, target_sheet, row_number, max_column):
    """Copy an entire row with values only"""
    for col in range(1, max_column + 1):
        source_cell = source_row[col - 1]
        new_cell = target_sheet.cell(row=row_number, column=col)
        new_cell.value = source_cell.value

        if source_cell.has_style:
            new_cell.font = source_cell.font.copy()
            new_cell.border = source_cell.border.copy()
            new_cell.fill = source_cell.fill.copy()
            new_cell.number_format = source_cell.number_format
            new_cell.protection = source_cell.protection.copy()
            new_cell.alignment = source_cell.alignment.copy()


def combine_excel_files(files):
    output = io.BytesIO()
    combined_wb = Workbook()
    combined_wb.remove(combined_wb.active)

    for file in files:
        wb = load_workbook(file, data_only=True)
        for sheet_name in wb.sheetnames:
            source_sheet = wb[sheet_name]
            new_sheet_name = f"{file.name.split('.')[0]}_{sheet_name}"[:31]
            target_sheet = combined_wb.create_sheet(title=new_sheet_name)

            max_column = source_sheet.max_column

            # Get header indices for all filters
            header_indices = []
            for filter_config in st.session_state.filters:
                if filter_config['enabled']:
                    header_idx = find_header_index(source_sheet, filter_config['header'])
                    header_indices.append(header_idx)

            # Copy headers (first row)
            first_row = next(source_sheet.iter_rows(min_row=1, max_row=1))
            copy_row_to_sheet(first_row, target_sheet, 1, max_column)

            # Copy filtered data
            new_row_idx = 2
            for row in source_sheet.iter_rows(min_row=2):
                if should_include_row(row, header_indices, st.session_state.filters):
                    copy_row_to_sheet(row, target_sheet, new_row_idx, max_column)
                    new_row_idx += 1

            # Copy styles and formatting
            wb_styles = load_workbook(file, data_only=False)
            source_sheet_styles = wb_styles[sheet_name]

            copy_column_dimensions(source_sheet_styles, target_sheet)
            copy_row_dimensions(source_sheet_styles, target_sheet)

            for merged_cell in source_sheet_styles.merged_cells.ranges:
                target_sheet.merge_cells(str(merged_cell))

            target_sheet.conditional_formatting = source_sheet_styles.conditional_formatting
            target_sheet.sheet_properties = source_sheet_styles.sheet_properties

    combined_wb.save(output)
    return output.getvalue()


def combine_into_single_sheet(files):
    output = io.BytesIO()
    combined_wb = Workbook()
    sheet = combined_wb.active
    sheet.title = "Combined"

    first_file = True
    current_row = 1
    max_column = 0

    for file in files:
        wb = load_workbook(file, data_only=True)
        source_sheet = wb.active

        max_column = max(max_column, source_sheet.max_column)

        # Get header indices for all filters
        header_indices = []
        for filter_config in st.session_state.filters:
            if filter_config['enabled']:
                header_idx = find_header_index(source_sheet, filter_config['header'])
                header_indices.append(header_idx)

        if first_file:
            first_row = next(source_sheet.iter_rows(min_row=1, max_row=1))
            copy_row_to_sheet(first_row, sheet, 1, max_column)

            wb_styles = load_workbook(file, data_only=False)
            source_sheet_styles = wb_styles.active
            copy_column_dimensions(source_sheet_styles, sheet)

            current_row = 2
            first_file = False

        # Copy filtered data
        for row in source_sheet.iter_rows(min_row=2):
            if should_include_row(row, header_indices, st.session_state.filters):
                copy_row_to_sheet(row, sheet, current_row, max_column)
                current_row += 1

    combined_wb.save(output)
    return output.getvalue()


def auto_download(bin_file, file_label='File'):
    b64 = base64.b64encode(bin_file).decode()
    return f"""
        <html>
        <body>
        <script>
            function download(filename, content) {{
                var element = document.createElement('a');
                element.setAttribute('href', 'data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,' + content);
                element.setAttribute('download', filename);
                element.style.display = 'none';
                document.body.appendChild(element);
                element.click();
                document.body.removeChild(element);
            }}
            download("{file_label}.xlsx", "{b64}");
        </script>
        </body>
        </html>
    """


# Add space above buttons
st.markdown("<br><br>", unsafe_allow_html=True)

# Button layout
col1, col2, col3 = st.columns([2, 3, 2])

with col2:
    button_col1, button_col2 = st.columns(2)

    with button_col1:
        if st.button("COMBINE IN DIFF SHEETS"):
            if uploaded_files:
                with st.spinner("Combining Excel files into different sheets..."):
                    combined_excel = combine_excel_files(uploaded_files)
                st.components.v1.html(auto_download(combined_excel, 'Combined_Excel_MultiSheet'), height=0)
            else:
                st.error("Please upload Excel files before combining.")

    with button_col2:
        if st.button("COMBINE IN SINGLE SHEET"):
            if uploaded_files:
                with st.spinner("Combining Excel files into a single sheet..."):
                    combined_excel = combine_into_single_sheet(uploaded_files)
                st.components.v1.html(auto_download(combined_excel, 'Combined_Excel_SingleSheet'), height=0)
            else:
                st.error("Please upload Excel files before combining.")